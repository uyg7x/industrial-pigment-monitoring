"""Export analysis results to a formatted multi-sheet Excel workbook."""

import os
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
_H_FILL   = PatternFill("solid", fgColor="1A3A5C")
_ALT_FILL = PatternFill("solid", fgColor="EBF0F7")
_GRN_FILL = PatternFill("solid", fgColor="D4EDDA")
_YEL_FILL = PatternFill("solid", fgColor="FFF3CD")
_RED_FILL = PatternFill("solid", fgColor="F8D7DA")
_CRT_FILL = PatternFill("solid", fgColor="721C24")

_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_WHITE_FONT = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
_BOLD_FONT  = Font(name="Calibri", bold=True, size=10)
_BODY_FONT  = Font(name="Calibri", size=10)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1A3A5C")


def _header_row(ws, row, headers):
    for c, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill      = _H_FILL
        cell.font      = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _THIN


def _data_rows(ws, start, data, fill_map=None):
    for r, row_vals in enumerate(data):
        alt = _ALT_FILL if r % 2 == 1 else None
        for c, val in enumerate(row_vals, 1):
            cell = ws.cell(row=start + r, column=c, value=val)
            cell.font      = _BODY_FONT
            cell.border    = _THIN
            cell.alignment = Alignment(vertical="center")
            if fill_map and c == fill_map.get("col"):
                fill = fill_map["fn"](val)
                if fill:
                    cell.fill = fill
            elif alt:
                cell.fill = alt


def _auto_width(ws):
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[
            get_column_letter(col[0].column)].width = min(max(width + 2, 10), 35)


def _title(ws, text):
    ws.merge_cells("A1:H1")
    ws["A1"].value     = text
    ws["A1"].font      = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font  = Font(name="Calibri", size=9, italic=True, color="888888")
    return 4


def _summary_sheet(wb, df, summary):
    ws = wb.create_sheet("Summary")
    row = _title(ws, "Pigment Process Monitor — Export Summary")
    kpis = [
        ("Total Batches",      summary["total_batches"]),
        ("Avg COD (mg/L)",     summary["avg_cod"]),
        ("Avg BOD (mg/L)",     summary["avg_bod"]),
        ("Equipment Failures", summary["fail_count"]),
        ("Failure Rate (%)",   f"{summary['fail_rate']}%"),
        ("Destinations",       summary["destinations"]),
        ("Total Quantity",     summary["total_quantity"]),
        ("Avg Vibration",      summary["avg_vibration"]),
        ("Avg Runtime (h)",    summary["avg_runtime"]),
        ("Avg Temperature",    summary["avg_temp"]),
        ("Avg pH",             summary["avg_ph"]),
        ("Avg Motor Current",  summary["avg_motor"]),
    ]
    _header_row(ws, row, ["KPI", "Value"])
    _data_rows(ws, row + 1, kpis)
    _auto_width(ws)
    ws.sheet_view.showGridLines = False


def _raw_sheet(wb, df):
    ws = wb.create_sheet("Raw Data")
    row = _title(ws, "Raw Batch Data")
    _header_row(ws, row, list(df.columns))
    data = [list(r) for r in df.itertuples(index=False, name=None)]
    _data_rows(ws, row + 1, data)
    _auto_width(ws)
    ws.sheet_view.showGridLines = False


def _cod_bod_sheet(wb, preds):
    ws = wb.create_sheet("COD-BOD Predictions")
    row = _title(ws, "COD / BOD — Actual vs Predicted")
    headers = list(preds.columns)
    _header_row(ws, row, headers)
    for r_idx, rec in enumerate(preds.itertuples(index=False, name=None)):
        for c_idx, val in enumerate(rec, 1):
            cell = ws.cell(row=row + 1 + r_idx, column=c_idx, value=val)
            cell.font = _BODY_FONT; cell.border = _THIN
            cell.alignment = Alignment(vertical="center")
            if r_idx % 2: cell.fill = _ALT_FILL
    _auto_width(ws)
    ws.sheet_view.showGridLines = False


def _colour_sheet(wb, preds):
    ws = wb.create_sheet("Colour Predictions")
    row = _title(ws, "Pigment Colour Analysis")
    headers = list(preds.columns)
    _header_row(ws, row, headers)
    fills = {"In-Spec": _GRN_FILL, "Marginal": _YEL_FILL, "Out-of-Spec": _RED_FILL}
    for r_idx, rec in enumerate(preds.itertuples(index=False, name=None)):
        rd = dict(zip(headers, rec))
        bf = fills.get(rd.get("Color_Status",""), _ALT_FILL if r_idx%2 else None)
        for c_idx, val in enumerate(rec, 1):
            cell = ws.cell(row=row+1+r_idx, column=c_idx, value=val)
            cell.font = _BODY_FONT; cell.border = _THIN
            cell.alignment = Alignment(vertical="center")
            if bf: cell.fill = bf
    _auto_width(ws)
    ws.sheet_view.showGridLines = False


def _equip_sheet(wb, preds):
    ws = wb.create_sheet("Equipment Health")
    row = _title(ws, "Equipment Failure Risk Scores")
    headers = list(preds.columns)
    _header_row(ws, row, headers)
    fills = {"Low":_GRN_FILL,"Medium":_YEL_FILL,"High":_RED_FILL}
    fonts = {"Critical": Font(name="Calibri", size=10, color="FFFFFF")}
    for r_idx, rec in enumerate(preds.itertuples(index=False, name=None)):
        rd  = dict(zip(headers, rec))
        lv  = rd.get("Risk_Level","")
        bf  = fills.get(lv, _CRT_FILL if lv=="Critical" else (_ALT_FILL if r_idx%2 else None))
        fnt = fonts.get(lv, _BODY_FONT)
        for c_idx, val in enumerate(rec, 1):
            cell = ws.cell(row=row+1+r_idx, column=c_idx, value=val)
            cell.font = fnt; cell.border = _THIN
            cell.alignment = Alignment(vertical="center")
            if bf: cell.fill = bf
    _auto_width(ws)
    ws.sheet_view.showGridLines = False


def _delivery_sheet(wb, df):
    ws = wb.create_sheet("Delivery Tracking")
    row = _title(ws, "Delivery Status by Batch")
    today  = pd.Timestamp.now().normalize()
    del_df = df[["Batch","Destination","DueDate","Quantity"]].copy()
    del_df["Days_Until_Due"] = (del_df["DueDate"] - today).dt.days
    del_df["Status"] = del_df["Days_Until_Due"].apply(
        lambda d: "Overdue" if pd.notna(d) and d < 0
        else ("Due Soon" if pd.notna(d) and d <= 3 else "On Track"))
    fills = {"On Track":_GRN_FILL,"Due Soon":_YEL_FILL,"Overdue":_RED_FILL}
    headers = list(del_df.columns)
    _header_row(ws, row, headers)
    for r_idx, rec in enumerate(del_df.itertuples(index=False, name=None)):
        rd = dict(zip(headers, rec))
        bf = fills.get(rd.get("Status",""), _ALT_FILL if r_idx%2 else None)
        for c_idx, val in enumerate(rec, 1):
            if hasattr(val, "strftime"): val = val.strftime("%Y-%m-%d")
            cell = ws.cell(row=row+1+r_idx, column=c_idx, value=val)
            cell.font = _BODY_FONT; cell.border = _THIN
            cell.alignment = Alignment(vertical="center")
            if bf: cell.fill = bf
    _auto_width(ws)
    ws.sheet_view.showGridLines = False


def export_to_excel(df, cod_bod_preds, color_preds, equip_preds, path):
    from app.data_loader import get_summary
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]
    summary = get_summary(df)
    _summary_sheet(wb, df, summary)
    _raw_sheet(wb, df)
    _cod_bod_sheet(wb, cod_bod_preds)
    _colour_sheet(wb, color_preds)
    _equip_sheet(wb, equip_preds)
    _delivery_sheet(wb, df)
    os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)
    wb.save(path)
